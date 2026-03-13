import openpyxl
import pandas as pd
import re
import json

excel_path = "_Tabelle sport di squadra_2026.xlsx"
index_path = "index.html"

wb = openpyxl.load_workbook(excel_path, data_only=True)
ws = wb["Tabelle attivita"]

def extract_table(ws, header_row, start_col, ncols, stop_row=None):
    headers=[ws.cell(header_row, start_col+i).value for i in range(ncols)]
    r=header_row+1
    rows=[]
    maxr=stop_row or ws.max_row
    blank_run=0

    while r<=maxr:
        vals=[ws.cell(r,start_col+i).value for i in range(ncols)]

        if all(v is None or (isinstance(v,str) and v.strip()=="") for v in vals):
            blank_run+=1
            if blank_run>=3:
                break
        else:
            blank_run=0
            rows.append(vals)

        r+=1

    return headers, rows


h1, rows1 = extract_table(ws,5,1,6)
h2, rows2 = extract_table(ws,5,8,6)
h3, rows3 = extract_table(ws,5,15,6)

df1=pd.DataFrame(rows1, columns=h1)
df2=pd.DataFrame(rows2, columns=h2)
df3=pd.DataFrame(rows3, columns=h3)

def clean(df, note=False):

    df=df.copy()

    for col in ["Disciplina","categoria","Regione"]:
        if col in df.columns:
            df[col]=df[col].ffill()

    if "Squadre" in df.columns:
        df["Squadre"]=pd.to_numeric(df["Squadre"], errors="coerce")

    if note==False:
        df["NOTE"]=""

    if "NOTE" not in df.columns:
        df["NOTE"]=""

    return df


df1=clean(df1)
df2=clean(df2,True)
df3=clean(df3,True)

def convert(df,status):

    rows=[]

    for _,r in df.iterrows():

        rows.append({

            "status":status,
            "Disciplina":r.get("Disciplina",""),
            "categoria":r.get("categoria",""),
            "Regione":r.get("Regione",""),
            "Comitato":r.get("Comitato",""),
            "Squadre":None if pd.isna(r.get("Squadre")) else int(r.get("Squadre")),
            "NOTE":r.get("NOTE","")

        })

    return rows


data = (

    convert(df1,"Soddisfa C.U.")
    + convert(df2,"non soddisfa c.u.")
    + convert(df3,"Non previsto nazionali")

)

new_data="const EMBEDDED_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";"

with open(index_path,"r",encoding="utf-8") as f:
    html=f.read()

pattern=r"const EMBEDDED_DATA\s*=\s*\[.*?\];"

html=re.sub(pattern,new_data,html,flags=re.S)

with open(index_path,"w",encoding="utf-8") as f:
    f.write(html)

print("index aggiornato")
